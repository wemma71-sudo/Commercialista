#!/usr/bin/env python3
"""
Web application per eseguire run_pipeline.py con gestione template Excel
Fornisce un'interfaccia web con selezione cartella e esecuzione pipeline
"""

import base64
import subprocess
import sys
from pathlib import Path
from flask import Flask, render_template, request, jsonify
import json
import os
from threading import Thread
import queue
from datetime import datetime
from shutil import copy2
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pyperclip
import webbrowser
import zipfile
import tempfile
import pandas as pd
import pythoncom
import re
import time
import urllib.request
import urllib.error
from dotenv import load_dotenv

# Moduli per l'invio al Sistema Tessera Sanitaria
import genera_invio_ts
import invia_ts
import ricevute_ts

# Carica le variabili d'ambiente dal file .env
load_dotenv()

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Queue per gestire output dei processi
process_queue = queue.Queue()
# Variabili globali per il risultato
execution_result = {}


@app.route('/')
def index():
    """Pagina principale"""
    ts_ambiente = os.getenv('TS_AMBIENTE', 'test').strip().lower()
    return render_template('index.html', ts_ambiente=ts_ambiente)


@app.route('/api/browse-folder', methods=['POST'])
def browse_folder():
    """
    Apre il dialog di selezione cartella di Windows
    Ritorna il percorso selezionato
    """
    try:
        import tkinter as tk
        from tkinter import filedialog
        
        folder_type = request.json.get('type', 'input')
        title_map = {
            'input': 'Seleziona cartella per l\'elaborazione',
            'template': 'Seleziona cartella con template',
            'private': 'Seleziona cartella di input fatture private'
        }
        
        # Crea una finestra Tkinter nascosta
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        # Apri dialog di selezione cartella
        folder_path = filedialog.askdirectory(
            title=title_map.get(folder_type, 'Seleziona cartella'),
            initialdir=os.path.expanduser("~\\Documents")
        )
        
        root.destroy()
        
        if folder_path:
            return jsonify({
                'success': True,
                'folder': folder_path
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Nessuna cartella selezionata'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400


@app.route('/api/run-pipeline', methods=['POST'])
def run_pipeline():
    """
    Esegue il workflow completo:
    1. Copia template Excel e lo rinomina con data
    2. Esegue run_pipeline.py
    3. Apre file Excel e incolla clipboard
    4. Salva il file
    """
    global execution_result
    execution_result = {}
    
    try:
        data = request.json
        input_folder = data.get('folder', '').strip()
        template_folder = data.get('template_folder', '').strip()
        
        if not input_folder:
            return jsonify({
                'success': False,
                'error': 'Cartella di input non specificata'
            }), 400
        
        if not template_folder:
            return jsonify({
                'success': False,
                'error': 'Cartella con template non specificata'
            }), 400
        
        # Verifica che le cartelle esistano
        if not os.path.isdir(input_folder):
            return jsonify({
                'success': False,
                'error': f'Cartella di input non trovata: {input_folder}'
            }), 400
        
        if not os.path.isdir(template_folder):
            return jsonify({
                'success': False,
                'error': f'Cartella template non trovata: {template_folder}'
            }), 400
        
        # Avvia il processo in background
        def run_process():
            global execution_result
            try:
                # Step 1: Trova e copia il template
                template_path = Path(template_folder) / 'template_ReportIncassiPagamenti.xlsx'
                
                if not template_path.exists():
                    execution_result = {
                        'status': 'error',
                        'error': f'Template non trovato: {template_path}',
                        'success': False
                    }
                    return
                
                # Genera nome file con data corrente
                today = datetime.now().strftime('%Y%m%d')
                output_filename = f'{today}_ReportIncassiPagamenti.xlsx'
                output_path = Path(template_folder) / output_filename
                
                # Copia il template
                copy2(str(template_path), str(output_path))
                
                # Step 2: Esegui run_pipeline.py
                result = subprocess.run(
                    [sys.executable, 'run_pipeline.py', input_folder],
                    cwd=Path(__file__).parent,
                    capture_output=True,
                    text=True
                )
                
                if result.returncode != 0:
                    execution_result = {
                        'status': 'error',
                        'error': f'Pipeline fallita: {result.stderr}',
                        'success': False,
                        'output_file': str(output_path)
                    }
                    return
                
                # Step 3: Apri file Excel e incolla clipboard
                try:
                    wb = load_workbook(str(output_path))
                    
                    # Accedi al foglio "Fatture Attive"
                    if 'Fatture Attive' not in wb.sheetnames:
                        execution_result = {
                            'status': 'error',
                            'error': f'Foglio "Fatture Attive" non trovato. Fogli disponibili: {wb.sheetnames}',
                            'success': False,
                            'output_file': str(output_path)
                        }
                        wb.close()
                        return
                    
                    ws = wb['Fatture Attive']
                    
                    # Copia il contenuto della clipboard
                    clipboard_content = pyperclip.paste()
                    
                    if not clipboard_content:
                        execution_result = {
                            'status': 'warning',
                            'error': 'Clipboard è vuoto, file creato senza dati',
                            'success': False,
                            'output_file': str(output_path)
                        }
                        wb.close()
                        return
                    
                    # Incolla il contenuto nella cella A2
                    # Se è una lista di righe, inseriamo i dati
                    lines = clipboard_content.strip().split('\n')
                    for row_idx, line in enumerate(lines, start=2):
                        # Se la riga contiene tab, è una tabella
                        if '\t' in line:
                            cells = line.split('\t')
                            for col_idx, cell_value in enumerate(cells, start=1):
                                ws.cell(row=row_idx, column=col_idx).value = cell_value
                        else:
                            ws.cell(row=row_idx, column=1).value = line
                    
                    # Step 4: Salva il file
                    wb.save(str(output_path))
                    wb.close()
                    
                    execution_result = {
                        'status': 'completed',
                        'success': True,
                        'message': f'Pipeline completata con successo',
                        'output_file': str(output_path),
                        'stdout': result.stdout,
                        'stderr': result.stderr
                    }
                    
                except Exception as e:
                    execution_result = {
                        'status': 'error',
                        'error': f'Errore durante l\'elaborazione Excel: {str(e)}',
                        'success': False,
                        'output_file': str(output_path)
                    }
                
            except Exception as e:
                execution_result = {
                    'status': 'error',
                    'error': str(e),
                    'success': False
                }
        
        # Avvia il processo in un thread separato
        thread = Thread(target=run_process, daemon=True)
        thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Pipeline in esecuzione...',
            'input_folder': input_folder,
            'template_folder': template_folder
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400


@app.route('/api/check-status', methods=['GET'])
def check_status():
    """
    Controlla lo stato della pipeline in esecuzione
    """
    global execution_result
    
    if execution_result and execution_result.get('status') in ['completed', 'error', 'warning']:
        result = execution_result.copy()
        execution_result = {}
        return jsonify(result)
    
    return jsonify({'status': 'running'})


@app.route('/api/open-file', methods=['POST'])
def open_file():
    """
    Apre il file Excel generato
    """
    try:
        data = request.json
        file_path = data.get('file_path', '').strip()
        
        if not file_path or not os.path.isfile(file_path):
            return jsonify({
                'success': False,
                'error': f'File non trovato: {file_path}'
            }), 400
        
        # Apri il file con l'applicazione predefinita (Excel)
        os.startfile(file_path)
        
        return jsonify({
            'success': True,
            'message': f'File aperto: {file_path}'
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400


def find_latest_zip(folder: Path) -> Path | None:
    """Trova l'ultimo file .zip nella cartella"""
    zips = list(folder.glob("*.zip"))
    if not zips:
        return None
    return max(zips, key=lambda p: p.stat().st_mtime)


def find_first_excel(root: Path) -> Path | None:
    """Trova il primo file Excel nella cartella (ricorsivo)"""
    for p in root.rglob('*'):
        if p.suffix.lower() in ('.xlsx', '.xlsm', '.xls') and p.is_file():
            return p
    return None


def move_excel_column(excel_path: Path, src_col: str = 'W', after_col: str = 'E') -> None:
    """Sposta una colonna Excel dopo un'altra usando win32com"""
    try:
        import win32com.client
    except ImportError as exc:
        raise ImportError('pywin32 is required. Install with: pip install pywin32') from exc

    pythoncom.CoInitialize()
    try:
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = None
        try:
            workbook = excel.Workbooks.Open(str(excel_path), False, False)
            if workbook is None:
                raise RuntimeError(f'Excel failed to open workbook: {excel_path}')
            sheet = workbook.Worksheets(1)
            src_range = f'{src_col}:{src_col}'
            insert_before_col = chr(ord(after_col.upper()) + 1)
            insert_range = f'{insert_before_col}:{insert_before_col}'
            xlToRight = -4161

            sheet.Columns(src_range).Cut()
            sheet.Columns(insert_range).Insert(Shift=xlToRight)

            workbook.Save()
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=True)
            excel.Quit()
    finally:
        pythoncom.CoUninitialize()


def copy_data_excluding_header_to_clipboard(excel_path: Path) -> None:
    """Copia i dati (escluso header) del primo foglio Excel in clipboard"""
    try:
        import win32com.client
    except ImportError as exc:
        raise ImportError('pywin32 is required. Install with: pip install pywin32') from exc

    pythoncom.CoInitialize()
    try:
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = None
        try:
            workbook = excel.Workbooks.Open(str(excel_path), False, False)
            sheet = workbook.Worksheets(1)
            used = sheet.UsedRange
            last_row = used.Rows.Count
            last_col = used.Columns.Count
            if last_row <= 1:
                return
            start = sheet.Cells(2, 1)
            end = sheet.Cells(last_row, last_col)
            rng = sheet.Range(start, end)
            rng.Copy()
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
            excel.Quit()
    finally:
        pythoncom.CoUninitialize()


def _importo_per_cella(valore):
    """Converte l'importo estratto in un numero per la cella Excel.

    Mantiene il formato numerico del template. Se il valore è 'NA' o non
    convertibile, restituisce la stringa originale (così 'NA' resta visibile).
    """
    if valore is None:
        return None
    s = str(valore).strip()
    if not s or s.upper() == 'NA':
        return s or None
    normalized = s.replace(' ', '').replace('€', '')
    if ',' in normalized and '.' in normalized:
        normalized = normalized.replace('.', '').replace(',', '.')
    elif ',' in normalized:
        normalized = normalized.replace(',', '.')
    try:
        f = float(normalized)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def estrai_campi_ricevuta_pdf(pdf_path: Path) -> dict:
    """Estrae data, numero, nome paziente e importo da una ricevuta PDF via Claude API.

    Restituisce sempre un dict con le chiavi 'data', 'numero', 'nome_cliente',
    'importo'. I campi non riconosciuti (o in caso di errore) valgono 'NA'.
    """
    default = {'data': 'NA', 'numero': 'NA', 'nome_cliente': 'NA', 'importo': 'NA'}

    # I PDF delle ricevute sono spesso scansioni/immagini: PyPDF2 non ne estrae
    # testo. Inviamo quindi il PDF stesso a Claude (content block "document",
    # base64), così il modello lo legge come farebbe l'interfaccia di claude.ai.
    try:
        pdf_b64 = base64.standard_b64encode(pdf_path.read_bytes()).decode('ascii')
    except Exception as e:
        return {**default, 'errore': f'Lettura PDF fallita: {e}'}

    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        return {**default, 'errore': 'API key non configurata'}

    system_prompt = """Sei un estrattore di informazioni da file PDF di ricevute/fatture mediche italiane.

Devi estrarre dal documento in input:
- data ricevuta (la data di emissione del documento)
- numero ricevuta
- nome e cognome del paziente
- importo ricevuta

Il medico è "Dott.ssa Isabella EUSEBIO" o simile: NON è il paziente, ignoralo.
Il paziente è la persona che ha ricevuto la prestazione sanitaria.

Se non riconosci uno dei campi, imposta il default "NA".

Rispondi SOLO con un oggetto JSON in questo formato esatto, senza testo prima o dopo:
{"data": "gg/mm/aaaa", "numero": "...", "nome_cliente": "Nome Cognome", "importo": "120,00"}

Regole di formato:
- data: formato gg/mm/aaaa
- importo: solo il numero con separatore decimale e senza simbolo € (es. "120,00")
- usa esattamente "NA" per ogni campo non determinabile."""

    payload = {
        "model": "claude-sonnet-4-6",
        "max_tokens": 300,
        "system": system_prompt,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64,
                        },
                    },
                    {"type": "text", "text": "Estrai i campi richiesti dalla ricevuta."},
                ],
            }
        ]
    }

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=data,
        headers={
            "Content-Type": "application/json",
            "anthropic-version": "2023-06-01",
            "x-api-key": api_key,
        },
        method="POST"
    )

    try:
        with urllib.request.urlopen(req) as resp:
            risposta = json.loads(resp.read().decode("utf-8"))
        testo_risposta = risposta["content"][0]["text"].strip()
        testo_risposta = re.sub(r"```(?:json)?", "", testo_risposta).strip()
        result = json.loads(testo_risposta)
    except urllib.error.HTTPError as e:
        # Mostra il corpo della risposta dell'API (es. modello inesistente, auth) invece del solo codice HTTP
        try:
            dettaglio = e.read().decode("utf-8")
        except Exception:
            dettaglio = ""
        return {**default, 'errore': f"HTTP {e.code}: {dettaglio}"}
    except Exception as e:
        return {**default, 'errore': str(e)}

    # Normalizza: ogni campo mancante/vuoto diventa "NA"
    campi = {}
    for chiave in ('data', 'numero', 'nome_cliente', 'importo'):
        valore = result.get(chiave)
        valore = str(valore).strip() if valore is not None else ''
        campi[chiave] = valore if valore else 'NA'

    # Normalizza il nome paziente: iniziale maiuscola di ogni parola, resto minuscolo.
    if campi['nome_cliente'] != 'NA':
        campi['nome_cliente'] = campi['nome_cliente'].title()

    return campi


@app.route('/api/process-private', methods=['POST'])
def process_private():
    """
    Elabora le fatture private dai file .txt e le scrive nel foglio "Fatture Privati" del file di output.
    """
    global execution_result
    execution_result = {}

    try:
        data = request.json
        private_folder = data.get('private_folder', '').strip()
        output_file = data.get('output_file', '').strip()

        if not private_folder:
            return jsonify({'success': False, 'error': 'Cartella di input fatture private non specificata'}), 400
        if not output_file:
            return jsonify({'success': False, 'error': 'File di output non specificato'}), 400
        if not os.path.isdir(private_folder):
            return jsonify({'success': False, 'error': f'Cartella di input fatture private non trovata: {private_folder}'}), 400
        if not os.path.isfile(output_file):
            return jsonify({'success': False, 'error': f'File di output non trovato: {output_file}'}), 400

        def run_process():
            global execution_result
            try:
                pdf_files = sorted([p for p in Path(private_folder).glob('*.pdf') if p.is_file()])
                if not pdf_files:
                    execution_result = {'status': 'error', 'error': f'Nessun file .pdf trovato in: {private_folder}', 'success': False}
                    return

                try:
                    output_wb = load_workbook(output_file)
                except Exception as e:
                    execution_result = {'status': 'error', 'error': f'Impossibile aprire file di output: {str(e)}', 'success': False}
                    return

                if 'Fatture Privati' not in output_wb.sheetnames:
                    output_wb.close()
                    execution_result = {
                        'status': 'error',
                        'error': f'Foglio "Fatture Privati" non trovato. Fogli disponibili: {output_wb.sheetnames}',
                        'success': False
                    }
                    return

                ws = output_wb['Fatture Privati']
                row = 2
                for pdf_file in pdf_files:
                    # Tutti i campi (data, numero, nome, importo) sono estratti dal
                    # PDF tramite Claude API; i campi non riconosciuti valgono 'NA'.
                    parsed = estrai_campi_ricevuta_pdf(pdf_file)
                    # L'importo va scritto come numero: così la cella conserva il
                    # formato presente sul template invece di essere trattato come testo.
                    importo_cell = _importo_per_cella(parsed.get('importo'))
                    ws.cell(row=row, column=1).value = 'FATTURA'
                    ws.cell(row=row, column=2).value = parsed.get('numero', 'NA')
                    ws.cell(row=row, column=3).value = parsed.get('data', 'NA')
                    ws.cell(row=row, column=4).value = importo_cell
                    ws.cell(row=row, column=5).value = parsed.get('data', 'NA')
                    ws.cell(row=row, column=6).value = parsed.get('nome_cliente', 'NA')
                    row += 1

                try:
                    output_wb.save(output_file)
                    output_wb.close()
                except Exception as e:
                    output_wb.close()
                    execution_result = {'status': 'error', 'error': f'Errore salvando il file di output: {str(e)}', 'success': False}
                    return

                execution_result = {'status': 'completed', 'success': True, 'message': 'Fatture private elaborate con successo', 'output_file': output_file}
            except Exception as e:
                execution_result = {'status': 'error', 'error': str(e), 'success': False}

        thread = Thread(target=run_process, daemon=True)
        thread.start()

        return jsonify({'success': True, 'message': 'Elaborazione fatture private in corso...'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/process-passive', methods=['POST'])
def process_passive():
    """
    Elabora le fatture passive:
    1. Trova l'ultimo .zip in input folder
    2. Unzippa e estrae Excel
    3. Sposta colonna W dopo E
    4. Copia dati (escluso header) in clipboard
    5. Apre il file output
    6. Incolla nel foglio "Fatture Passive" a partire da A2
    7. Salva il file
    """
    global execution_result
    execution_result = {}
    
    try:
        data = request.json
        input_folder = data.get('input_folder', '').strip()
        output_file = data.get('output_file', '').strip()
        
        if not input_folder:
            return jsonify({
                'success': False,
                'error': 'Cartella di input non specificata'
            }), 400
        
        if not output_file:
            return jsonify({
                'success': False,
                'error': 'File di output non specificato'
            }), 400
        
        # Verifica che le cartelle/file esistano
        if not os.path.isdir(input_folder):
            return jsonify({
                'success': False,
                'error': f'Cartella di input non trovata: {input_folder}'
            }), 400
        
        if not os.path.isfile(output_file):
            return jsonify({
                'success': False,
                'error': f'File di output non trovato: {output_file}'
            }), 400
        
        # Avvia il processo in background
        def run_process():
            global execution_result
            try:
                # Step 1: Trova l'ultimo file .zip
                input_path = Path(input_folder)
                zip_file = find_latest_zip(input_path)
                
                if not zip_file:
                    execution_result = {
                        'status': 'error',
                        'error': f'Nessun file .zip trovato in: {input_folder}',
                        'success': False
                    }
                    return
                
                # Step 2: Unzippa il file
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_path = Path(temp_dir)
                    with zipfile.ZipFile(zip_file, 'r') as z:
                        z.extractall(temp_path)
                    
                    # Step 3: Trova il file Excel estratto
                    excel_file = find_first_excel(temp_path)
                    if not excel_file:
                        execution_result = {
                            'status': 'error',
                            'error': f'Nessun file Excel trovato nel .zip: {zip_file}',
                            'success': False
                        }
                        return
                    
                    # Step 4: Sposta colonna W dopo E
                    try:
                        move_excel_column(excel_file, src_col='W', after_col='E')
                    except Exception as e:
                        execution_result = {
                            'status': 'error',
                            'error': f'Errore durante lo spostamento colonna: {str(e)}',
                            'success': False
                        }
                        return
                    
                    # Step 5: Copia dati in clipboard
                    try:
                        copy_data_excluding_header_to_clipboard(excel_file)
                    except Exception as e:
                        execution_result = {
                            'status': 'error',
                            'error': f'Errore durante la copia in clipboard: {str(e)}',
                            'success': False
                        }
                        return
                    
                    # Step 6: Apri il file output e incolla nel foglio "Fatture Passive"
                    try:
                        output_wb = load_workbook(output_file)
                        
                        if 'Fatture Passive' not in output_wb.sheetnames:
                            execution_result = {
                                'status': 'error',
                                'error': f'Foglio "Fatture Passive" non trovato. Fogli disponibili: {output_wb.sheetnames}',
                                'success': False
                            }
                            output_wb.close()
                            return
                        
                        ws = output_wb['Fatture Passive']
                        
                        # Leggi il file Excel estratto usando pandas (supporta sia .xls che .xlsx)
                        try:
                            df = pd.read_excel(excel_file, sheet_name=0)
                        except Exception as e:
                            execution_result = {
                                'status': 'error',
                                'error': f'Errore nella lettura del file Excel: {str(e)}',
                                'success': False
                            }
                            output_wb.close()
                            return
                        
                        # Scrivi i dati nel foglio "Fatture Passive" a partire da A2
                        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                            for col_idx, (_, value) in enumerate(row.items(), start=1):
                                ws.cell(row=row_idx, column=col_idx).value = value
                        
                        # Step 7: Salva il file
                        output_wb.save(output_file)
                        output_wb.close()
                        
                        execution_result = {
                            'status': 'completed',
                            'success': True,
                            'message': f'Fatture passive elaborate con successo',
                            'output_file': output_file
                        }
                        
                    except Exception as e:
                        execution_result = {
                            'status': 'error',
                            'error': f'Errore durante l\'elaborazione del file output: {str(e)}',
                            'success': False
                        }
                
            except Exception as e:
                execution_result = {
                    'status': 'error',
                    'error': str(e),
                    'success': False
                }
        
        # Avvia il processo in un thread separato
        thread = Thread(target=run_process, daemon=True)
        thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Elaborazione fatture passive in corso...'
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400


# Partita IVA fissa dello studio, usata in tutti i record del foglio documentoSpesa
DOCUMENTO_SPESA_PIVA = 'IT03972130045'
CODICE_FISCALE_PROPRIETARIO = 'SBESLL71B46B111Q'


@app.route('/api/output-ts', methods=['POST'])
def output_ts():
    """
    Genera i record del foglio "documentoSpesa" a partire dalle fatture
    presenti nel foglio "Fatture Privati" del file di output.

    Per ogni fattura privata (a partire dalla riga 2) crea un record nel
    foglio "documentoSpesa" a partire dalla riga 2 con la struttura prevista
    dal tracciato Sistema TS.
    """
    global execution_result
    execution_result = {}

    try:
        data = request.json
        output_file = data.get('output_file', '').strip()

        if not output_file:
            return jsonify({'success': False, 'error': 'File di output non specificato'}), 400
        if not os.path.isfile(output_file):
            return jsonify({'success': False, 'error': f'File di output non trovato: {output_file}'}), 400

        def run_process():
            global execution_result
            try:
                try:
                    output_wb = load_workbook(output_file)
                except Exception as e:
                    execution_result = {'status': 'error', 'error': f'Impossibile aprire file di output: {str(e)}', 'success': False}
                    return

                if 'Fatture Privati' not in output_wb.sheetnames:
                    output_wb.close()
                    execution_result = {
                        'status': 'error',
                        'error': f'Foglio "Fatture Privati" non trovato. Fogli disponibili: {output_wb.sheetnames}',
                        'success': False
                    }
                    return

                if 'documentoSpesa' not in output_wb.sheetnames:
                    output_wb.close()
                    execution_result = {
                        'status': 'error',
                        'error': f'Foglio "documentoSpesa" non trovato. Fogli disponibili: {output_wb.sheetnames}',
                        'success': False
                    }
                    return

                src = output_wb['Fatture Privati']
                dst = output_wb['documentoSpesa']

                dst_row = 2
                count = 0
                # Le fatture private partono dalla riga 2; la colonna A contiene
                # sempre il marker 'FATTURA'. Si scorre finché si trovano dati.
                src_row = 2
                while True:
                    marker = src.cell(row=src_row, column=1).value          # A
                    numero = src.cell(row=src_row, column=2).value          # B -> NumDocumento
                    data_emissione = src.cell(row=src_row, column=3).value  # C -> dataEmissione
                    importo = src.cell(row=src_row, column=4).value        # D -> importo
                    data_pagamento = src.cell(row=src_row, column=5).value  # E -> dataPagamento
                    cf_cittadino = src.cell(row=src_row, column=7).value    # G -> cfCittadino

                    # Riga vuota -> fine dei dati
                    if not marker and not numero and not data_emissione:
                        break
                    dst.cell(row=dst_row, column=1).value = CODICE_FISCALE_PROPRIETARIO  # codice fiscale    
                    dst.cell(row=dst_row, column=2).value = DOCUMENTO_SPESA_PIVA  # pIVA
                    dst.cell(row=dst_row, column=3).value = data_emissione        # dataEmissione
                    dst.cell(row=dst_row, column=4).value = 1                     # dispositivo
                    dst.cell(row=dst_row, column=5).value = numero                # NumDocumento
                    dst.cell(row=dst_row, column=6).value = data_pagamento        # dataPagamento
                    dst.cell(row=dst_row, column=7).value = None                  # flagPagamentoAnticipato
                    dst.cell(row=dst_row, column=8).value = 'I'                   # flagOperazione
                    dst.cell(row=dst_row, column=9).value = cf_cittadino          # cfCittadino
                    dst.cell(row=dst_row, column=10).value = 'SI'                  # pagamentoTracciato
                    dst.cell(row=dst_row, column=11).value = 'F'                  # tipoDocumento
                    dst.cell(row=dst_row, column=12).value = 0                    # flagOpposizione
                    dst.cell(row=dst_row, column=13).value = 'SR'                 # tipoSpesa
                    dst.cell(row=dst_row, column=14).value = None                 # flagTipoSpesa
                    dst.cell(row=dst_row, column=15).value = None                 # AliquotaIva
                    dst.cell(row=dst_row, column=16).value = 'N4'                 # naturaIVA
                    dst.cell(row=dst_row, column=17).value = None                 # idRimborso
                    dst.cell(row=dst_row, column=18).value = importo              # R -> importo (da Fatture Privati col. D)

                    dst_row += 1
                    src_row += 1
                    count += 1

                try:
                    output_wb.save(output_file)
                    output_wb.close()
                except Exception as e:
                    output_wb.close()
                    execution_result = {'status': 'error', 'error': f'Errore salvando il file di output: {str(e)}', 'success': False}
                    return

                execution_result = {
                    'status': 'completed',
                    'success': True,
                    'message': f'Output TS generato: {count} record creati nel foglio "documentoSpesa"',
                    'output_file': output_file
                }
            except Exception as e:
                execution_result = {'status': 'error', 'error': str(e), 'success': False}

        thread = Thread(target=run_process, daemon=True)
        thread.start()

        return jsonify({'success': True, 'message': 'Generazione Output TS in corso...'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/invia-ts', methods=['POST'])
def invia_ts_route():
    """
    Invio completo al Sistema Tessera Sanitaria:
    1. Genera l'XML <precompilata> + .zip dal foglio "documentoSpesa" del file di output
    2. Trasmette il file (SOAP MTOM) all'ambiente configurato (TS_AMBIENTE)
    3. Recupera l'esito e scarica la ricevuta PDF

    Le credenziali di produzione sono lette dal file .env:
    TS_AMBIENTE (test|produzione), TS_USERNAME, TS_PASSWORD, TS_PINCODE, TS_CF_PROPRIETARIO
    """
    global execution_result
    execution_result = {}

    try:
        data = request.json
        output_file = data.get('output_file', '').strip()
        richiesta_ambiente = (data.get('ambiente') or '').strip()

        if not output_file:
            return jsonify({'success': False, 'error': 'File di output non specificato'}), 400
        if not os.path.isfile(output_file):
            return jsonify({'success': False, 'error': f'File di output non trovato: {output_file}'}), 400

        def run_process():
            global execution_result
            try:
                # L'ambiente può essere scelto dal frontend; in mancanza usa quello del .env
                ambiente = (richiesta_ambiente or os.getenv('TS_AMBIENTE', 'test')).strip().lower()
                prod = ambiente.startswith('prod')

                # Step 1: genera XML + ZIP dal foglio documentoSpesa
                try:
                    gen = genera_invio_ts.genera_file(output_file)
                except Exception as e:
                    execution_result = {'status': 'error', 'success': False,
                                        'error': f'Errore nella generazione del file: {str(e)}',
                                        'output_file': output_file}
                    return
                zip_path = gen['zip_path']

                # Credenziali e endpoint in base all'ambiente
                if prod:
                    user = os.getenv('TS_USERNAME', '').strip()
                    password = os.getenv('TS_PASSWORD', '').strip()
                    pincode = os.getenv('TS_PINCODE', '').strip()
                    cf_prop = os.getenv('TS_CF_PROPRIETARIO', '').strip()
                    mancanti = [n for n, v in (('TS_USERNAME', user), ('TS_PASSWORD', password),
                                               ('TS_PINCODE', pincode), ('TS_CF_PROPRIETARIO', cf_prop)) if not v]
                    if mancanti:
                        execution_result = {'status': 'error', 'success': False,
                                            'error': f'Credenziali di produzione mancanti nel file .env: {", ".join(mancanti)}',
                                            'output_file': output_file}
                        return
                    endpoint = invia_ts.ENDPOINT_PROD
                    verify = True
                else:
                    user, password = invia_ts.TEST_USER, invia_ts.TEST_PASSWORD
                    pincode, cf_prop = invia_ts.TEST_PINCODE, invia_ts.TEST_CF_PROPRIETARIO
                    endpoint = invia_ts.ENDPOINT_TEST
                    verify = False

                # Step 2: invio del file
                try:
                    campi = invia_ts.invia(endpoint, user, password, pincode, cf_prop, zip_path, verify=verify)
                except Exception as e:
                    execution_result = {'status': 'error', 'success': False,
                                        'error': f'Errore durante l\'invio: {str(e)}',
                                        'output_file': output_file}
                    return

                codice = campi.get('codiceEsito')
                protocollo = campi.get('protocollo')
                descrizione_invio = campi.get('descrizioneEsito', '')

                if codice != '000' or not protocollo:
                    execution_result = {'status': 'error', 'success': False,
                                        'error': f'Invio non accolto (codice {codice}): {descrizione_invio}',
                                        'ambiente': ambiente, 'documenti': gen['n'],
                                        'output_file': output_file}
                    return

                # Step 3: attende l'elaborazione e recupera l'esito
                esito = {}
                for _ in range(6):
                    time.sleep(5)
                    try:
                        esito = ricevute_ts.recupera_esito(protocollo, user, password, pincode, prod=prod, verify=verify)
                    except Exception:
                        esito = {}
                    if esito.get('stato'):  # esito di merito disponibile
                        break

                # Scarica la ricevuta PDF (disponibile anche in caso di scarto)
                ricevuta_file = None
                try:
                    out_pdf = Path(output_file).parent / f'ricevuta_{protocollo}.pdf'
                    got = ricevute_ts.scarica_ricevuta_pdf(protocollo, out_pdf, user, password, pincode, prod=prod, verify=verify)
                    if got:
                        ricevuta_file = str(got)
                except Exception:
                    ricevuta_file = None

                stato = esito.get('stato')
                accolto = stato == '2'
                if accolto:
                    msg = (f'Invio ACCOLTO ({ambiente}). Protocollo {protocollo}. '
                           f'Documenti accolti: {esito.get("nAccolti", "?")} su {esito.get("nInviati", "?")}, '
                           f'errori: {esito.get("nErrori", "0")}.')
                else:
                    descr = esito.get('descrizione') or descrizione_invio or 'in elaborazione'
                    msg = (f'Invio trasmesso ({ambiente}), protocollo {protocollo}. '
                           f'Esito: {descr}. Apri la ricevuta per i dettagli.')

                execution_result = {
                    'status': 'completed',
                    'success': bool(accolto),
                    'message': msg,
                    'protocollo': protocollo,
                    'ambiente': ambiente,
                    'documenti': gen['n'],
                    'esito': esito,
                    'ricevuta_file': ricevuta_file,
                    'output_file': output_file,
                }
            except Exception as e:
                execution_result = {'status': 'error', 'success': False, 'error': str(e),
                                    'output_file': output_file}

        thread = Thread(target=run_process, daemon=True)
        thread.start()

        return jsonify({'success': True, 'message': 'Invio al Sistema TS in corso...'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


if __name__ == '__main__':
    print("Applicazione avviata su http://localhost:5000")
    print("Premi Ctrl+C per fermare l'applicazione")
    app.run(debug=False, host='localhost', port=5000)
