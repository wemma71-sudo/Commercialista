#!/usr/bin/env python3
"""
Genera il file di invio per il Sistema Tessera Sanitaria (730 spese sanitarie,
modalità asincrona, caso d'uso Professionista).

Legge il foglio "documentoSpesa" del file Excel di output (e l'importo dal foglio
"Fatture Privati"), costruisce l'XML conforme allo schema ufficiale
ts730/730_precompilata.xsd, cifra i codici fiscali con il certificato pubblico
SanitelCF (ts730/SanitelCF.cer) e produce sia l'XML sia lo .zip da allegare.

Riferimenti ufficiali (Development Kit kit730P_ver_20240214):
- schema:      ts730/730_precompilata.xsd  (root <precompilata>)
- certificato: ts730/SanitelCF.cer         (RSA 1024, valido 23/01/2024 -> 23/01/2027)
- esempio:     ts730/esempio_medico.xml

Uso:
    python genera_invio_ts.py "C:\\percorso\\20260609_ReportIncassiPagamenti.xlsx"

Se il percorso non è indicato, usa il file di esempio in OneDrive.
"""

import sys
import zipfile
from base64 import b64encode
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.x509 import load_der_x509_certificate, load_pem_x509_certificate

try:
    from lxml import etree
except ImportError:
    etree = None

# --- Percorsi delle risorse ufficiali (cartella ts730 accanto a questo script) ---
BASE_DIR = Path(__file__).parent
CERT_PATH = BASE_DIR / 'ts730' / 'SanitelCF.cer'
XSD_PATH = BASE_DIR / 'ts730' / '730_precompilata.xsd'

# Indici di colonna (1-based) del foglio "documentoSpesa"
COL = {
    'cfProprietario': 1,   # A
    'pIVA': 2,             # B
    'dataEmissione': 3,    # C
    'dispositivo': 4,      # D
    'numDocumento': 5,     # E
    'dataPagamento': 6,    # F
    'flagPagamentoAnticipato': 7,  # G
    'flagOperazione': 8,   # H
    'cfCittadino': 9,      # I
    'pagamentoTracciato': 10,  # J
    'tipoDocumento': 11,   # K
    'flagOpposizione': 12, # L
    'tipoSpesa': 13,       # M
    'flagTipoSpesa': 14,   # N
    'aliquotaIVA': 15,     # O
    'naturaIVA': 16,       # P
    'idRimborso': 17,      # Q
    'importo': 18,         # R (popolato da Output TS dalla col. D di Fatture Privati)
}


def carica_certificato(path: Path):
    """Carica il certificato SanitelCF (prova DER, poi PEM) e ne restituisce la chiave pubblica RSA."""
    raw = path.read_bytes()
    try:
        cert = load_der_x509_certificate(raw)
    except Exception:
        cert = load_pem_x509_certificate(raw)
    try:
        inizio, fine = cert.not_valid_before_utc, cert.not_valid_after_utc
        oggi = datetime.now(inizio.tzinfo)
    except AttributeError:  # cryptography < 42
        inizio, fine = cert.not_valid_before, cert.not_valid_after
        oggi = datetime.now()
    if not (inizio <= oggi <= fine):
        print(f'ATTENZIONE: il certificato SanitelCF non è valido oggi '
              f'(valido {inizio:%d/%m/%Y} - {fine:%d/%m/%Y}). '
              f'Scaricane uno aggiornato dal Development Kit del portale sistemats.it.')
    return cert.public_key()


def cifra_cf(public_key, cf: str) -> str:
    """Cifra un codice fiscale con la chiave pubblica (RSA PKCS#1 v1.5) e lo codifica in base64."""
    cf = (cf or '').strip().upper()
    cifrato = public_key.encrypt(cf.encode('utf-8'), padding.PKCS1v15())
    return b64encode(cifrato).decode('ascii')


def to_iso_date(valore) -> str:
    """Converte una data ('dd/mm/yyyy', 'yyyy-mm-dd' o datetime) nel formato xs:date 'yyyy-mm-dd'."""
    if valore is None or valore == '':
        raise ValueError('data mancante')
    if isinstance(valore, datetime):
        return valore.strftime('%Y-%m-%d')
    s = str(valore).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    raise ValueError(f'formato data non riconosciuto: {valore!r}')


def normalizza_piva(valore) -> str:
    """Estrae le 11 cifre della partita IVA, eliminando il prefisso 'IT' o altri caratteri."""
    cifre = ''.join(ch for ch in str(valore) if ch.isdigit())
    if len(cifre) != 11:
        raise ValueError(f'partita IVA non valida (attese 11 cifre): {valore!r} -> {cifre!r}')
    return cifre


def normalizza_dispositivo(valore) -> int:
    """Il dispositivo vale sempre 1 per fatture/ricevute; corregge celle errate (es. formattate come data)."""
    try:
        n = int(valore)
        if 1 <= n <= 999:
            return n
    except (TypeError, ValueError):
        pass
    return 1


def formatta_importo(valore) -> str:
    """Formatta l'importo come richiesto dallo schema: 1-5 cifre intere e 2 decimali (es. '120.00')."""
    if valore is None or valore == '':
        raise ValueError('importo mancante')
    num = float(str(valore).replace(',', '.'))
    if num < 0.01:
        raise ValueError(f'importo non valido (minimo 0.01): {valore!r}')
    if num > 99999.99:
        raise ValueError(f'importo fuori range (max 99999.99): {valore!r}')
    return f'{num:.2f}'


def mappa_importi_da_privati(wb) -> dict:
    """Costruisce {numero_documento: importo} dal foglio 'Fatture Privati' (col. B=Numero, D=Importo)."""
    if 'Fatture Privati' not in wb.sheetnames:
        return {}
    ws = wb['Fatture Privati']
    importi = {}
    for r in range(2, ws.max_row + 1):
        numero = ws.cell(row=r, column=2).value
        importo = ws.cell(row=r, column=4).value
        if numero is None:
            continue
        importi[str(numero).strip()] = importo
    return importi


def costruisci_xml(righe: list, cf_proprietario_cifrato: str) -> str:
    """Costruisce il contenuto XML <precompilata> a partire dalle righe già elaborate."""
    out = []
    out.append('<?xml version="1.0" encoding="UTF-8"?>')
    out.append('<precompilata xsi:noNamespaceSchemaLocation="730_precompilata.xsd" '
               'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">')
    out.append('\t<proprietario>')
    out.append(f'\t\t<cfProprietario>{cf_proprietario_cifrato}</cfProprietario>')
    out.append('\t</proprietario>')

    for r in righe:
        out.append('\t<documentoSpesa>')
        out.append('\t\t<idSpesa>')
        out.append(f'\t\t\t<pIva>{r["pIva"]}</pIva>')
        out.append(f'\t\t\t<dataEmissione>{r["dataEmissione"]}</dataEmissione>')
        out.append('\t\t\t<numDocumentoFiscale>')
        out.append(f'\t\t\t\t<dispositivo>{r["dispositivo"]}</dispositivo>')
        out.append(f'\t\t\t\t<numDocumento>{escape(r["numDocumento"])}</numDocumento>')
        out.append('\t\t\t</numDocumentoFiscale>')
        out.append('\t\t</idSpesa>')
        out.append(f'\t\t<dataPagamento>{r["dataPagamento"]}</dataPagamento>')
        if r.get('flagPagamentoAnticipato'):
            out.append('\t\t<flagPagamentoAnticipato>1</flagPagamentoAnticipato>')
        out.append(f'\t\t<flagOperazione>{r["flagOperazione"]}</flagOperazione>')
        # cfCittadino deve essere ASSENTE se flagOpposizione = 1
        if r.get('cfCittadino') and r.get('flagOpposizione') != '1':
            out.append(f'\t\t<cfCittadino>{r["cfCittadino"]}</cfCittadino>')
        if r.get('pagamentoTracciato'):
            out.append(f'\t\t<pagamentoTracciato>{r["pagamentoTracciato"]}</pagamentoTracciato>')
        if r.get('tipoDocumento'):
            out.append(f'\t\t<tipoDocumento>{r["tipoDocumento"]}</tipoDocumento>')
        if r.get('flagOpposizione') is not None:
            out.append(f'\t\t<flagOpposizione>{r["flagOpposizione"]}</flagOpposizione>')
        out.append('\t\t<voceSpesa>')
        out.append(f'\t\t\t<tipoSpesa>{r["tipoSpesa"]}</tipoSpesa>')
        if r.get('flagTipoSpesa'):
            out.append(f'\t\t\t<flagTipoSpesa>{r["flagTipoSpesa"]}</flagTipoSpesa>')
        out.append(f'\t\t\t<importo>{r["importo"]}</importo>')
        if r.get('aliquotaIVA'):
            out.append(f'\t\t\t<aliquotaIVA>{r["aliquotaIVA"]}</aliquotaIVA>')
        elif r.get('naturaIVA'):
            out.append(f'\t\t\t<naturaIVA>{r["naturaIVA"]}</naturaIVA>')
        out.append('\t\t</voceSpesa>')
        out.append('\t</documentoSpesa>')

    out.append('</precompilata>')
    return '\n'.join(out)


def elabora(excel_path: Path):
    wb = load_workbook(excel_path, data_only=True)
    if 'documentoSpesa' not in wb.sheetnames:
        raise ValueError(f'Foglio "documentoSpesa" non trovato. Fogli: {wb.sheetnames}')
    ws = wb['documentoSpesa']
    importi = mappa_importi_da_privati(wb)

    public_key = carica_certificato(CERT_PATH)

    cf_proprietario_plain = None
    righe = []
    for r in range(2, ws.max_row + 1):
        def cell(nome):
            v = ws.cell(row=r, column=COL[nome]).value
            return v.strip() if isinstance(v, str) else v

        # Fine dati: riga senza pIVA e senza numero documento
        if not cell('pIVA') and not cell('numDocumento') and not cell('cfProprietario'):
            continue

        if cf_proprietario_plain is None and cell('cfProprietario'):
            cf_proprietario_plain = str(cell('cfProprietario')).strip().upper()

        num_doc = str(cell('numDocumento')).strip()
        # importo: preferisci la colonna R di documentoSpesa, altrimenti recupera da Fatture Privati
        importo_raw = cell('importo')
        if importo_raw in (None, ''):
            importo_raw = importi.get(num_doc)
        flag_opp = cell('flagOpposizione')
        flag_opp = str(int(flag_opp)) if isinstance(flag_opp, (int, float)) else (str(flag_opp).strip() if flag_opp is not None else None)

        aliquota = cell('aliquotaIVA')
        natura = cell('naturaIVA')

        riga = {
            'pIva': normalizza_piva(cell('pIVA')),
            'dataEmissione': to_iso_date(cell('dataEmissione')),
            'dispositivo': normalizza_dispositivo(cell('dispositivo')),
            'numDocumento': num_doc,
            'dataPagamento': to_iso_date(cell('dataPagamento')),
            'flagOperazione': (str(cell('flagOperazione')).strip().upper() or 'I'),
            'cfCittadino': cifra_cf(public_key, str(cell('cfCittadino'))) if cell('cfCittadino') else None,
            'pagamentoTracciato': (str(cell('pagamentoTracciato')).strip().upper() if cell('pagamentoTracciato') else None),
            'tipoDocumento': (str(cell('tipoDocumento')).strip().upper() if cell('tipoDocumento') else None),
            'flagOpposizione': flag_opp,
            'tipoSpesa': str(cell('tipoSpesa')).strip().upper(),
            'flagTipoSpesa': (str(cell('flagTipoSpesa')).strip() if cell('flagTipoSpesa') else None),
            'importo': formatta_importo(importo_raw),
            'aliquotaIVA': (f'{float(str(aliquota).replace(",", ".")):.2f}' if aliquota not in (None, '') else None),
            'naturaIVA': (str(natura).strip().upper() if natura else None),
        }
        # pagamento anticipato se la data di pagamento è precedente all'emissione
        if riga['dataPagamento'] < riga['dataEmissione']:
            riga['flagPagamentoAnticipato'] = True
        righe.append(riga)

    if not righe:
        raise ValueError('Nessuna riga trovata nel foglio "documentoSpesa".')
    if not cf_proprietario_plain:
        raise ValueError('cfProprietario non valorizzato nel foglio "documentoSpesa".')

    cf_prop_cifrato = cifra_cf(public_key, cf_proprietario_plain)
    xml = costruisci_xml(righe, cf_prop_cifrato)
    return xml, len(righe)


def valida(xml: str):
    """Valida l'XML contro lo schema ufficiale. Ritorna (ok: bool, errori: list[str])."""
    if etree is None:
        return True, ['lxml non disponibile: validazione XSD saltata.']
    schema = etree.XMLSchema(etree.parse(str(XSD_PATH)))
    doc = etree.fromstring(xml.encode('utf-8'))
    if schema.validate(doc):
        return True, []
    return False, [f'riga {e.line}: {e.message}' for e in schema.error_log]


def genera_file(excel_path, out_dir=None):
    """
    Elabora l'Excel, valida e scrive XML + ZIP. Ritorna un dict con i percorsi
    e l'esito della validazione. Solleva ValueError sui problemi di dati.
    """
    excel_path = Path(excel_path)
    out_dir = Path(out_dir) if out_dir else excel_path.parent
    xml, n = elabora(excel_path)
    ok, errori = valida(xml)
    if not ok:
        raise ValueError('XML non valido secondo lo schema TS:\n' + '\n'.join(errori))

    oggi = datetime.now().strftime('%Y%m%d')
    base = f'{oggi}_invio730'
    xml_path = out_dir / f'{base}.xml'
    zip_path = out_dir / f'{base}.zip'
    xml_path.write_text(xml, encoding='utf-8')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
        z.write(xml_path, arcname=xml_path.name)
    return {'xml_path': xml_path, 'zip_path': zip_path, 'n': n, 'valido': ok, 'note': errori}


def main():
    if len(sys.argv) > 1:
        excel_path = Path(sys.argv[1])
    else:
        excel_path = Path(r'C:\Users\francesco.emma\OneDrive - KERING SA\Walter'
                          r'\Contabilita Isabella\20260609_ReportIncassiPagamenti.xlsx')

    if not excel_path.is_file():
        raise SystemExit(f'File Excel non trovato: {excel_path}')
    if not CERT_PATH.is_file():
        raise SystemExit(f'Certificato non trovato: {CERT_PATH}')

    res = genera_file(excel_path)
    if res['note']:
        print('Validazione XSD:', '; '.join(res['note']))
    else:
        print('Validazione XSD: OK')
    print(f'Documenti elaborati: {res["n"]}')
    print(f'XML generato: {res["xml_path"]}')
    print(f'ZIP da allegare: {res["zip_path"]}')


if __name__ == '__main__':
    main()
